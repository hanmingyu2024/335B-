from __future__ import annotations

import argparse
import csv
import json
import re
from dataclasses import asdict, dataclass
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_WORKSPACE = ROOT / "artifacts" / "standard_line_workspace"

WORKBOOK_PATTERNS = (
    "3358B*/04_*/*.xlsx",
    "*/04_*/*.xlsx",
    "**/*IO*.xlsx",
    "**/*io*.xlsx",
)

AXIS_REQUIREMENTS = {
    "Station03": {
        "axis_name": "claw_horizontal_axis",
        "required_points": [
            {"direction": "Input", "address": "I0.0", "data_type": "Bool", "name": "home_sensor"},
            {"direction": "Input", "address": "I0.1", "data_type": "Bool", "name": "right_limit"},
            {"direction": "Input", "address": "I0.2", "data_type": "Bool", "name": "left_limit"},
            {"direction": "Output", "address": "Q0.0", "data_type": "Bool", "name": "axis_pulse"},
            {"direction": "Output", "address": "Q0.1", "data_type": "Bool", "name": "axis_direction"},
        ],
    },
    "Station05": {
        "axis_name": "conveyor_axis",
        "required_points": [
            {"direction": "Input", "address": "I0.0", "data_type": "Bool", "name": "encoder_b"},
            {"direction": "Input", "address": "I0.1", "data_type": "Bool", "name": "encoder_a"},
            {"direction": "Input", "address": "I0.2", "data_type": "Bool", "name": "encoder_z"},
            {"direction": "Output", "address": "Q0.0", "data_type": "Bool", "name": "motor_start"},
            {"direction": "Output", "address": "QW80", "data_type": "Word", "name": "analog_speed_ref"},
        ],
    },
}


@dataclass
class IoPoint:
    station_key: str
    station_name: str
    direction: str
    index: int
    symbol_name: str
    address: str
    data_type: str
    comment: str


def read_text_utf8(path: Path) -> str:
    return path.read_text(encoding="utf-8")


def write_text_utf8(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content.replace("\r\n", "\n").replace("\n", "\r\n"), encoding="utf-8")


def load_workspace_manifest(workspace_root: Path) -> dict:
    manifest_path = workspace_root / "config" / "line_manifest.json"
    return json.loads(read_text_utf8(manifest_path))


def discover_workbook(search_root: Path = ROOT) -> Path:
    candidates: list[Path] = []
    seen: set[str] = set()
    for pattern in WORKBOOK_PATTERNS:
        for candidate in sorted(search_root.glob(pattern)):
            if candidate.name.startswith("~$"):
                continue
            resolved = str(candidate.resolve()).lower()
            if resolved in seen:
                continue
            seen.add(resolved)
            candidates.append(candidate.resolve())
    if not candidates:
        raise FileNotFoundError(
            "No workbook found. Provide --workbook-path or place workbook under a '04_*' folder."
        )
    return candidates[0]


def station_prefix(name: str) -> str:
    match = re.match(r"^(\d{2})", name.strip())
    return match.group(1) if match else ""


def normalize_address(raw: str) -> str:
    value = str(raw or "").strip().replace("%", "").replace(" ", "")
    return value.upper()


def infer_type(address: str) -> str:
    if "." in address:
        return "Bool"
    if address.startswith(("IW", "QW")):
        return "Word"
    if address.startswith(("ID", "QD")):
        return "DWord"
    return "Bool"


def classify_prefix(direction: str, data_type: str) -> str:
    if direction == "Input":
        return "DI" if data_type == "Bool" else "AI"
    return "DO" if data_type == "Bool" else "AQ"


def build_symbol_name(station_key: str, direction: str, data_type: str, seq: int) -> str:
    prefix = classify_prefix(direction, data_type)
    return f"{station_key}_{prefix}_{seq:03d}"


def parse_sheet_points(sheet, station_key: str, station_name: str) -> list[IoPoint]:
    points: list[IoPoint] = []
    input_seq = 1
    output_seq = 1

    for row in sheet.iter_rows(min_row=5, values_only=True):
        in_name = row[1]
        in_address = row[2]
        out_name = row[5]
        out_address = row[6]

        if in_name and in_address:
            address = normalize_address(str(in_address))
            data_type = infer_type(address)
            points.append(
                IoPoint(
                    station_key=station_key,
                    station_name=station_name,
                    direction="Input",
                    index=input_seq,
                    symbol_name=build_symbol_name(station_key, "Input", data_type, input_seq),
                    address=address,
                    data_type=data_type,
                    comment=str(in_name).strip(),
                )
            )
            input_seq += 1

        if out_name and out_address:
            address = normalize_address(str(out_address))
            data_type = infer_type(address)
            points.append(
                IoPoint(
                    station_key=station_key,
                    station_name=station_name,
                    direction="Output",
                    index=output_seq,
                    symbol_name=build_symbol_name(station_key, "Output", data_type, output_seq),
                    address=address,
                    data_type=data_type,
                    comment=str(out_name).strip(),
                )
            )
            output_seq += 1

    return points


def build_station_lookup(manifest: dict) -> dict[str, dict]:
    lookup: dict[str, dict] = {}
    for station in manifest["stations"]:
        lookup[f"{int(station['stationNumber']):02d}"] = station
    return lookup


def parse_workbook(workbook_path: Path, manifest: dict) -> list[IoPoint]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    lookup = build_station_lookup(manifest)
    all_points: list[IoPoint] = []

    for sheet in workbook.worksheets:
        prefix = station_prefix(sheet.title)
        if not prefix or prefix not in lookup:
            continue
        station = lookup[prefix]
        all_points.extend(parse_sheet_points(sheet, station["key"], station["displayName"]))

    return all_points


def write_flat_csv(points: list[IoPoint], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "station_key",
                "station_name",
                "direction",
                "index",
                "symbol_name",
                "address",
                "data_type",
                "comment",
            ]
        )
        for point in points:
            writer.writerow(
                [
                    point.station_key,
                    point.station_name,
                    point.direction,
                    point.index,
                    point.symbol_name,
                    point.address,
                    point.data_type,
                    point.comment,
                ]
            )


def write_station_csv(points: list[IoPoint], path: Path) -> None:
    write_flat_csv(points, path)


def build_db_io(points: list[IoPoint]) -> str:
    declarations: list[str] = []
    for point in points:
        declarations.append(f"      {point.symbol_name} : {point.data_type};")

    body = "\n".join(declarations) if declarations else "      // No IO points parsed."
    return (
        'DATA_BLOCK "DB_IO"\n'
        "{ S7_Optimized_Access := 'TRUE' }\n"
        "VERSION : 0.1\n"
        "NON_RETAIN\n"
        "   VAR\n"
        f"{body}\n"
        "   END_VAR\n"
        "BEGIN\n"
        "END_DATA_BLOCK\n"
    )


def build_i_map(points: list[IoPoint]) -> str:
    lines = [
        'FUNCTION "FC_IMap" : Void',
        "VERSION : 0.1",
        "BEGIN",
    ]
    for point in points:
        if point.direction != "Input":
            continue
        lines.append(f'   "DB_IO".{point.symbol_name} := {point.address}; // {point.comment}')
    if len(lines) == 3:
        lines.append("   // No input points parsed.")
    lines.append("END_FUNCTION")
    return "\n".join(lines) + "\n"


def build_q_map(points: list[IoPoint]) -> str:
    lines = [
        'FUNCTION "FC_QMap" : Void',
        "VERSION : 0.1",
        "BEGIN",
    ]
    for point in points:
        if point.direction != "Output":
            continue
        lines.append(f'   {point.address} := "DB_IO".{point.symbol_name}; // {point.comment}')
    if len(lines) == 3:
        lines.append("   // No output points parsed.")
    lines.append("END_FUNCTION")
    return "\n".join(lines) + "\n"


def write_station_markdown(station: dict, points: list[IoPoint], path: Path) -> None:
    lines = [
        f"# {station['displayName']} IO Details",
        "",
        "| Direction | Symbol | Address | Type | Comment |",
        "|---|---|---|---|---|",
    ]
    for point in points:
        lines.append(
            f"| {point.direction} | {point.symbol_name} | {point.address} | {point.data_type} | {point.comment} |"
        )
    write_text_utf8(path, "\n".join(lines) + "\n")


def write_generated_json(points: list[IoPoint], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = [asdict(point) for point in points]
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def update_tia_manifest(tia_manifest_path: Path, grouped_points: dict[str, list[IoPoint]]) -> None:
    manifest = json.loads(read_text_utf8(tia_manifest_path))
    existing = {(entry["StationKey"], entry["RelativePath"]): entry for entry in manifest["Sources"]}

    for station in manifest["Stations"]:
        station_key = station["StationKey"]
        if station_key not in grouped_points:
            continue

        record_key = (station_key, r"DB\DB_IO.db")
        if record_key not in existing:
            manifest["Sources"].append(
                {
                    "StationKey": station_key,
                    "Order": 95,
                    "Category": "DB",
                    "RelativePath": r"DB\DB_IO.db",
                    "FileName": "DB_IO.db",
                }
            )

    manifest["Sources"] = sorted(
        manifest["Sources"],
        key=lambda entry: (entry["StationKey"], int(entry["Order"]), entry["RelativePath"]),
    )
    tia_manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")


def check_duplicate_points(points: list[IoPoint]) -> list[str]:
    errors: list[str] = []
    seen_symbols: set[str] = set()
    seen_addresses: set[tuple[str, str, str]] = set()
    for point in points:
        symbol_key = point.symbol_name
        if symbol_key in seen_symbols:
            errors.append(f"Duplicate symbol_name detected: {symbol_key}")
        else:
            seen_symbols.add(symbol_key)

        address_key = (point.station_key, point.direction, point.address)
        if address_key in seen_addresses:
            errors.append(
                f"Duplicate address in station {point.station_key}: {point.direction} {point.address}"
            )
        else:
            seen_addresses.add(address_key)
    return errors


def validate_manifest_io_counts(
    manifest: dict,
    grouped_points: dict[str, list[IoPoint]],
    allow_io_mismatch: bool,
) -> tuple[list[str], list[dict]]:
    mismatches: list[str] = []
    report: list[dict] = []

    for station in manifest["stations"]:
        station_key = station["key"]
        station_points = grouped_points.get(station_key, [])
        actual_inputs = sum(1 for p in station_points if p.direction == "Input")
        actual_outputs = sum(1 for p in station_points if p.direction == "Output")
        expected_inputs = int(station["io"]["inputs"])
        expected_outputs = int(station["io"]["outputs"])

        input_match = actual_inputs == expected_inputs
        output_match = actual_outputs == expected_outputs
        report.append(
            {
                "station_key": station_key,
                "expected_inputs": expected_inputs,
                "actual_inputs": actual_inputs,
                "expected_outputs": expected_outputs,
                "actual_outputs": actual_outputs,
                "match": input_match and output_match,
            }
        )

        if not input_match or not output_match:
            mismatches.append(
                f"{station_key}: manifest DI/DO={expected_inputs}/{expected_outputs}, workbook DI/DO={actual_inputs}/{actual_outputs}"
            )

    if mismatches and not allow_io_mismatch:
        return mismatches, report
    return [], report


def validate_axis_requirements(grouped_points: dict[str, list[IoPoint]]) -> tuple[list[str], list[dict]]:
    errors: list[str] = []
    report: list[dict] = []

    for station_key, requirement in AXIS_REQUIREMENTS.items():
        station_points = grouped_points.get(station_key, [])
        index = {(p.direction, p.address): p for p in station_points}

        station_report = {
            "station_key": station_key,
            "axis_name": requirement["axis_name"],
            "missing_points": [],
            "type_mismatch_points": [],
        }

        for required_point in requirement["required_points"]:
            key = (required_point["direction"], required_point["address"])
            actual = index.get(key)
            if actual is None:
                station_report["missing_points"].append(required_point)
                errors.append(
                    f"{station_key} missing axis point: {required_point['direction']} {required_point['address']} ({required_point['name']})"
                )
                continue
            if actual.data_type != required_point["data_type"]:
                station_report["type_mismatch_points"].append(
                    {
                        "name": required_point["name"],
                        "direction": required_point["direction"],
                        "address": required_point["address"],
                        "expected_data_type": required_point["data_type"],
                        "actual_data_type": actual.data_type,
                    }
                )
                errors.append(
                    f"{station_key} axis point type mismatch: {required_point['direction']} {required_point['address']} expected {required_point['data_type']} got {actual.data_type}"
                )

        report.append(station_report)

    return errors, report


def write_json(path: Path, payload: object) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workspace-root", default=str(DEFAULT_WORKSPACE))
    parser.add_argument("--workbook-path", default="")
    parser.add_argument("--allow-io-mismatch", action="store_true")
    parser.add_argument("--consistency-report-path", default="")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    workspace_root = Path(args.workspace_root).resolve()
    manifest = load_workspace_manifest(workspace_root)
    workbook_path = Path(args.workbook_path).resolve() if args.workbook_path else discover_workbook()

    points = parse_workbook(workbook_path, manifest)
    grouped: dict[str, list[IoPoint]] = {}
    for point in points:
        grouped.setdefault(point.station_key, []).append(point)

    duplicate_errors = check_duplicate_points(points)
    io_mismatch_errors, io_count_report = validate_manifest_io_counts(
        manifest, grouped, allow_io_mismatch=args.allow_io_mismatch
    )
    axis_errors, axis_report = validate_axis_requirements(grouped)

    consistency_report = {
        "workspace_root": str(workspace_root),
        "workbook_path": str(workbook_path),
        "allow_io_mismatch": bool(args.allow_io_mismatch),
        "io_count_check": io_count_report,
        "axis_check": axis_report,
        "duplicate_errors": duplicate_errors,
        "io_mismatch_errors": io_mismatch_errors,
        "axis_errors": axis_errors,
    }

    consistency_report_path = (
        Path(args.consistency_report_path).resolve()
        if args.consistency_report_path
        else workspace_root / "config" / "io_consistency_report.json"
    )
    write_json(consistency_report_path, consistency_report)

    blocking_errors = duplicate_errors + axis_errors + io_mismatch_errors
    if blocking_errors:
        joined = "\n".join(f"- {item}" for item in blocking_errors)
        raise RuntimeError(
            "I/O consistency validation failed.\n"
            f"Workbook: {workbook_path}\n"
            f"Report: {consistency_report_path}\n"
            f"{joined}"
        )

    io_root = workspace_root / "04_IOList"
    tia_root = workspace_root / "tia_sources"

    write_flat_csv(points, io_root / "plc_tags_flat.csv")
    write_generated_json(points, workspace_root / "config" / "io_points.generated.json")

    for station in manifest["stations"]:
        station_key = station["key"]
        station_points = grouped.get(station_key, [])
        write_station_csv(station_points, io_root / f"{station_key}_plc_tags.csv")
        write_station_markdown(station, station_points, io_root / f"{station_key}_io_points.md")

        station_root = tia_root / station_key
        write_text_utf8(station_root / "DB" / "DB_IO.db", build_db_io(station_points))
        write_text_utf8(station_root / "Blocks" / "10_IO" / "FC_IMap.scl", build_i_map(station_points))
        write_text_utf8(station_root / "Blocks" / "10_IO" / "FC_QMap.scl", build_q_map(station_points))

    update_tia_manifest(tia_root / "manifest.json", grouped)

    print(io_root / "plc_tags_flat.csv")
    print(workspace_root / "config" / "io_points.generated.json")
    print(consistency_report_path)


if __name__ == "__main__":
    main()
