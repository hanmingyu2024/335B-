from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


def u(text: str) -> str:
    return text.encode("ascii").decode("unicode_escape")


ROOT = Path(__file__).resolve().parent.parent
OUTPUT_DIR = ROOT / "output" / "spreadsheet"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

SOURCE = next(ROOT.parent.glob("*TIA Openness API*.csv"))


def load_source() -> pd.DataFrame:
    raw = pd.read_csv(SOURCE, encoding="utf-8-sig")
    raw.columns = ["col1", "col2", "col3", "col4", "col5"]

    df = raw.iloc[1:].copy()
    df.columns = ["seq", "module", "category", "detail", "scenario"]
    df["seq"] = df["seq"].ffill().astype(str)
    df["module"] = df["module"].ffill()
    return df.reset_index(drop=True)


def build_mapping() -> dict[str, tuple[str, str, str, str, str]]:
    return {
        "1": (
            u("\\u9879\\u76ee\\u6846\\u67b6\\u642d\\u5efa / \\u4f18\\u5316"),
            u("\\u9879\\u76ee\\u521b\\u5efa\\u3001\\u7ed3\\u6784\\u751f\\u6210\\u3001\\u4fdd\\u5b58\\u5f52\\u6863\\u3001\\u7248\\u672c\\u8ffd\\u8e2a"),
            u("\\u5df2\\u90e8\\u5206\\u8986\\u76d6"),
            u("\\u9ad8"),
            u("\\u4fdd\\u6301\\u6e05\\u5355\\u9a71\\u52a8\\uff0c\\u4f18\\u5148\\u56de\\u5199\\u6a21\\u677f\\u548c\\u811a\\u672c"),
        ),
        "2": (
            u("\\u8bbe\\u5907\\u7ec4\\u6001"),
            u("CPU/\\u8bbe\\u5907\\u521b\\u5efa\\u3001\\u540e\\u7eed\\u6269\\u5c55 IO \\u548c\\u7f51\\u7edc\\u914d\\u7f6e"),
            u("\\u5df2\\u90e8\\u5206\\u8986\\u76d6"),
            u("\\u9ad8"),
            u("\\u8865\\u7f51\\u7edc\\u3001\\u5b50\\u7f51\\u3001\\u6269\\u5c55\\u6a21\\u5757\\u548c\\u8bbe\\u5907\\u53c2\\u6570\\u81ea\\u52a8\\u5316"),
        ),
        "3": (
            u("\\u5de5\\u827a\\u6d41\\u7a0b / \\u7f16\\u7a0b"),
            u("\\u63e1\\u624b\\u5b57\\u6bb5\\u3001\\u8fde\\u63a5\\u5bf9\\u8c61\\u3001\\u901a\\u8baf DB \\u89c4\\u5212"),
            u("\\u5f85\\u6269\\u5c55"),
            u("\\u9ad8"),
            u("\\u8865 S7 \\u8fde\\u63a5\\u3001OPC UA\\u3001Modbus \\u548c\\u8d85\\u65f6\\u53c2\\u6570\\u751f\\u6210"),
        ),
        "4": (
            u("\\u9879\\u76ee\\u6846\\u67b6\\u642d\\u5efa / \\u7f16\\u7a0b"),
            u("OB/FB/FC/DB/UDT \\u751f\\u6210\\u3001\\u4e3b\\u5faa\\u73af\\u8c03\\u7528\\u94fe"),
            u("\\u5df2\\u90e8\\u5206\\u8986\\u76d6"),
            u("\\u9ad8"),
            u("\\u7ee7\\u7eed\\u52a0\\u5f3a\\u6a21\\u677f\\u5757\\u751f\\u6210\\u548c\\u5b9e\\u4f8b\\u5316\\u7b56\\u7565"),
        ),
        "5": (
            u("\\u8c03\\u8bd5"),
            u("\\u81ea\\u52a8\\u7f16\\u8bd1\\u3001\\u65e5\\u5fd7\\u89e3\\u6790\\u3001\\u95ee\\u9898\\u5b9a\\u4f4d"),
            u("\\u5df2\\u8986\\u76d6"),
            u("\\u9ad8"),
            u("\\u7ee7\\u7eed\\u6269\\u5c55\\u89c4\\u8303\\u68c0\\u67e5\\u548c\\u9759\\u6001\\u626b\\u63cf"),
        ),
        "6": (
            u("\\u8c03\\u8bd5"),
            u("\\u4e0b\\u8f7d\\u3001\\u4e0a\\u4f20\\u3001\\u5728\\u7ebf\\u8bca\\u65ad\\u3001\\u5f3a\\u5236\\u548c\\u76d1\\u63a7"),
            u("\\u5f85\\u6269\\u5c55"),
            u("\\u4e2d"),
            u("\\u8865\\u73b0\\u573a\\u4e0b\\u8f7d\\u3001\\u4e0a\\u4f20\\u548c\\u5728\\u7ebf\\u8bca\\u65ad\\u5165\\u53e3"),
        ),
        "7": (
            u("\\u4f18\\u5316 / \\u591a\\u9879\\u76ee\\u590d\\u7528"),
            u("\\u57fa\\u4e8e\\u6e05\\u5355\\u6279\\u91cf\\u751f\\u6210\\u4ea7\\u7ebf\\u9879\\u76ee"),
            u("\\u5df2\\u90e8\\u5206\\u8986\\u76d6"),
            u("\\u9ad8"),
            u("\\u4fdd\\u6301\\u6e05\\u5355\\u9a71\\u52a8\\uff0c\\u4f18\\u5148\\u56de\\u5199\\u6a21\\u677f\\u548c\\u811a\\u672c"),
        ),
    }


def enrich(df: pd.DataFrame) -> pd.DataFrame:
    mapping = build_mapping()
    mapped = [mapping.get(value, ("", "", "", "", "")) for value in df["seq"]]
    df = df.copy()
    df["template_stage"] = [value[0] for value in mapped]
    df["automation_landing"] = [value[1] for value in mapped]
    df["repo_status"] = [value[2] for value in mapped]
    df["priority"] = [value[3] for value in mapped]
    df["next_action"] = [value[4] for value in mapped]
    return df


def build_summary(df: pd.DataFrame) -> pd.DataFrame:
    summary = (
        df.groupby(
            ["seq", "module", "template_stage", "repo_status", "priority"],
            as_index=False,
        )
        .agg(
            {
                "detail": "count",
                "automation_landing": "first",
                "next_action": "first",
            }
        )
        .rename(columns={"detail": "capability_count"})
    )
    return summary


def style_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    for worksheet in workbook.worksheets:
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        for column_cells in worksheet.columns:
            sample = column_cells[:50]
            width = max(len(str(cell.value or "")) for cell in sample)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(
                max(width + 2, 12),
                40,
            )

    workbook.save(path)


def main() -> None:
    df = enrich(load_source())
    summary = build_summary(df)

    csv_path = OUTPUT_DIR / "tia_openness_v17_capability_matrix.csv"
    xlsx_path = OUTPUT_DIR / "tia_openness_v17_capability_matrix.xlsx"

    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="capability_matrix")
        summary.to_excel(writer, index=False, sheet_name="summary")

    style_workbook(xlsx_path)

    print(csv_path)
    print(xlsx_path)


if __name__ == "__main__":
    main()
